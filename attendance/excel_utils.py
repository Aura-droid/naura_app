from collections import Counter
from io import BytesIO
from pathlib import Path

from django.utils import timezone

from .models import ResultEntry, ResultStudent, ResultTemplateSubject


META_HEADERS = {
    "CENTRE NO.",
    "CANDIDATES NO.",
    "STUDENT NAME",
    "SEX",
    "POINTS",
    "DIV",
    "DIVISION",
    "INC",
    "ABS",
    "TOT",
    "TOTAL",
    "AVERAGE",
    "GPA",
}


def load_openpyxl():
    try:
        from openpyxl import load_workbook
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise RuntimeError(
            "openpyxl is required for Excel template import/export. "
            "Install it with `pip install openpyxl`."
        ) from exc
    return load_workbook, get_column_letter


def _clean_value(value):
    if value is None:
        return ""
    return str(value).strip()


def detect_header_row(worksheet):
    for row_number in range(1, min(worksheet.max_row, 40) + 1):
        values = {
            _clean_value(worksheet.cell(row=row_number, column=col).value).upper()
            for col in range(1, min(worksheet.max_column, 90) + 1)
        }
        if "STUDENT NAME" in values and "CANDIDATES NO." in values:
            return row_number
    raise ValueError("Could not locate the student header row in the uploaded workbook.")


def _value_kind(value):
    text = _clean_value(value)
    if not text:
        return "blank"
    if text == "-":
        return "dash"
    upper = text.upper()
    if upper in {"ABS", "INC"}:
        return "code"
    try:
        float(text)
        return "numeric"
    except ValueError:
        pass
    if upper.isalpha() and len(upper) <= 3:
        return "grade"
    return "text"


def detect_mark_columns(worksheet, header_row, first_student_row):
    _, get_column_letter = load_openpyxl()
    subjects = []
    sample_end_row = min(worksheet.max_row, first_student_row + 24)

    for column_index in range(1, worksheet.max_column + 1):
        header = _clean_value(worksheet.cell(row=header_row, column=column_index).value)
        upper_header = header.upper()
        if not header or upper_header in META_HEADERS or upper_header.isdigit():
            continue

        sample_values = [
            _clean_value(worksheet.cell(row=row_number, column=column_index).value)
            for row_number in range(first_student_row, sample_end_row + 1)
        ]
        kinds = Counter(_value_kind(value) for value in sample_values if value != "")
        if not kinds:
            continue

        numeric_like = kinds["numeric"] + kinds["dash"] + kinds["code"]
        grade_like = kinds["grade"]

        if numeric_like == 0:
            continue
        if grade_like > numeric_like:
            continue

        subjects.append(
            {
                "name": header,
                "column_index": column_index,
                "column_letter": get_column_letter(column_index),
            }
        )

    if not subjects:
        raise ValueError("No score columns were detected in the uploaded workbook.")

    return subjects


def detect_students(worksheet, first_student_row):
    students = []
    row_number = first_student_row

    while row_number <= worksheet.max_row:
        candidate_no = _clean_value(worksheet.cell(row=row_number, column=2).value)
        student_name = _clean_value(worksheet.cell(row=row_number, column=3).value)

        if candidate_no.isdigit() and student_name:
            students.append(
                {
                    "row_number": row_number,
                    "centre_no": _clean_value(worksheet.cell(row=row_number, column=1).value),
                    "candidate_no": candidate_no,
                    "student_name": student_name,
                    "sex": _clean_value(worksheet.cell(row=row_number, column=4).value),
                }
            )
            row_number += 1
            continue

        if students:
            break
        row_number += 1

    if not students:
        raise ValueError("No student rows were detected in the uploaded workbook.")

    return students


def rebuild_template_structure(template):
    load_workbook, _ = load_openpyxl()
    workbook = load_workbook(template.workbook.path)
    worksheet = workbook[workbook.sheetnames[0]]

    header_row = detect_header_row(worksheet)
    first_student_row = header_row + 1
    students = detect_students(worksheet, first_student_row)
    subjects = detect_mark_columns(worksheet, header_row, first_student_row)

    template.sheet_name = worksheet.title
    template.header_row = header_row
    template.first_student_row = students[0]["row_number"]
    template.last_student_row = students[-1]["row_number"]
    template.last_processed_at = timezone.now()
    template.processing_error = ""
    template.save(
        update_fields=[
            "sheet_name",
            "header_row",
            "first_student_row",
            "last_student_row",
            "last_processed_at",
            "processing_error",
            "updated_at",
        ]
    )

    template.subjects.all().delete()
    template.students.all().delete()
    template.entries.all().delete()

    ResultTemplateSubject.objects.bulk_create(
        [
            ResultTemplateSubject(
                template=template,
                name=subject["name"],
                column_letter=subject["column_letter"],
                column_index=subject["column_index"],
                display_order=index,
            )
            for index, subject in enumerate(subjects, start=1)
        ]
    )

    ResultStudent.objects.bulk_create(
        [
            ResultStudent(
                template=template,
                row_number=student["row_number"],
                centre_no=student["centre_no"],
                candidate_no=student["candidate_no"],
                student_name=student["student_name"],
                sex=student["sex"],
            )
            for student in students
        ]
    )


def build_export_workbook(template):
    load_workbook, _ = load_openpyxl()
    workbook = load_workbook(template.workbook.path)
    worksheet = workbook[template.sheet_name or workbook.sheetnames[0]]

    entries = ResultEntry.objects.filter(template=template).select_related("subject", "student")
    for entry in entries:
        cell = worksheet.cell(
            row=entry.student.row_number,
            column=entry.subject.column_index,
        )
        value = _clean_value(entry.raw_score)
        if not value:
            cell.value = "-"
            continue

        try:
            numeric = float(value)
        except ValueError:
            cell.value = value.upper()
        else:
            if numeric.is_integer():
                cell.value = int(numeric)
            else:
                cell.value = numeric

    calc_props = getattr(workbook, "calculation", None)
    if calc_props is not None:
        setattr(calc_props, "fullCalcOnLoad", True)
        setattr(calc_props, "forceFullCalc", True)

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def export_filename_for_template(template):
    safe_name = "".join(
        char if char.isalnum() or char in {"-", "_"} else "_"
        for char in template.name.strip().replace(" ", "_")
    )
    suffix = timezone.now().strftime("%Y%m%d_%H%M")
    original_name = Path(template.workbook.name).name
    if original_name.lower().endswith(".xlsx"):
        return f"{safe_name}_{suffix}.xlsx"
    return f"{safe_name}_{suffix}_{original_name}"
